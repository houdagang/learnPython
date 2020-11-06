import openpyxl
import datetime
import math

work_days=[ '2020-10-09', '2020-10-10', '2020-10-12',
            '2020-10-13', '2020-10-14', '2020-10-15', '2020-10-16', '2020-10-19', '2020-10-20',
            '2020-10-21', '2020-10-22', '2020-10-23', '2020-10-26', '2020-10-27', '2020-10-28',
            '2020-10-29', '2020-10-30','2020-11-02', '2020-11-03', '2020-11-04', '2020-11-05',
    '2020-11-06', '2020-11-09', '2020-11-10', '2020-11-11', '2020-11-12', '2020-11-13',
    '2020-11-16', '2020-11-17', '2020-11-18', '2020-11-19', '2020-11-20', '2020-11-23',
    '2020-11-24', '2020-11-25', '2020-11-26', '2020-11-27', '2020-11-30', '2020-12-01',
    '2020-12-02', '2020-12-03', '2020-12-04', '2020-12-07', '2020-12-08', '2020-12-09',
    '2020-12-10', '2020-12-11', '2020-12-14', '2020-12-15', '2020-12-16', '2020-12-17',
    '2020-12-18', '2020-12-21', '2020-12-22', '2020-12-23', '2020-12-24', '2020-12-25',
    '2020-12-28', '2020-12-29', '2020-12-30', '2020-12-31', '2021-01-04', '2021-01-05', '2021-01-06', '2021-01-07',
    '2021-01-08', '2021-01-11', '2021-01-12', '2021-01-13', '2021-01-14', '2021-01-15', '2021-01-18', '2021-01-19',
           '2021-01-20', '2021-01-21', '2021-01-22', '2021-01-25', '2021-01-26', '2021-01-27', '2021-01-28', '2021-01-29',
           '2021-02-01', '2021-02-02', '2021-02-03', '2021-02-04', '2021-02-05', '2021-02-08', '2021-02-09', '2021-02-10',
           '2021-02-18', '2021-02-19', '2021-02-20', '2021-02-21', '2021-02-22', '2021-02-23', '2021-02-24', '2021-02-25',
           '2021-02-26', '2021-03-01', '2021-03-02', '2021-03-03', '2021-03-04', '2021-03-05', '2021-03-08', '2021-03-09',
           '2021-03-10', '2021-03-11', '2021-03-12', '2021-03-15', '2021-03-16', '2021-03-17', '2021-03-18', '2021-03-19',
           '2021-03-22', '2021-03-23', '2021-03-24', '2021-03-25', '2021-03-26', '2021-03-29', '2021-03-30', '2021-03-31',
           '2021-04-01', '2021-04-02', '2021-04-06', '2021-04-07', '2021-04-08', '2021-04-09', '2021-04-12', '2021-04-13',
           '2021-04-14', '2021-04-15', '2021-04-16', '2021-04-19', '2021-04-20', '2021-04-21', '2021-04-22', '2021-04-23',
           '2021-04-26', '2021-04-27', '2021-04-28', '2021-04-29', '2021-04-30', '2021-05-04', '2021-05-05', '2021-05-06',
           '2021-05-07', '2021-05-10', '2021-05-11', '2021-05-12', '2021-05-13', '2021-05-14', '2021-05-17', '2021-05-18',
           '2021-05-19', '2021-05-20', '2021-05-21', '2021-05-24', '2021-05-25', '2021-05-26', '2021-05-27', '2021-05-28',
           '2021-05-31', '2021-06-01', '2021-06-02', '2021-06-03', '2021-06-04', '2021-06-07', '2021-06-08', '2021-06-09',
           '2021-06-10', '2021-06-11', '2021-06-15', '2021-06-16', '2021-06-17', '2021-06-18', '2021-06-21', '2021-06-22',
           '2021-06-23', '2021-06-24', '2021-06-25', '2021-06-28', '2021-06-29', '2021-06-30', '2021-07-01', '2021-07-02',
           '2021-07-05', '2021-07-06', '2021-07-07', '2021-07-08', '2021-07-09', '2021-07-12', '2021-07-13', '2021-07-14',
           '2021-07-15', '2021-07-16', '2021-07-19', '2021-07-20', '2021-07-21', '2021-07-22', '2021-07-23', '2021-07-26',
           '2021-07-27', '2021-07-28', '2021-07-29', '2021-07-30', '2021-08-02', '2021-08-03', '2021-08-04', '2021-08-05',
           '2021-08-06', '2021-08-09', '2021-08-10', '2021-08-11', '2021-08-12', '2021-08-13', '2021-08-16', '2021-08-17',
           '2021-08-18', '2021-08-19', '2021-08-20', '2021-08-23', '2021-08-24', '2021-08-25', '2021-08-26', '2021-08-27',
           '2021-08-30', '2021-08-31', '2021-09-01', '2021-09-02', '2021-09-03', '2021-09-06', '2021-09-07', '2021-09-08',
           '2021-09-09', '2021-09-10', '2021-09-13', '2021-09-14', '2021-09-15', '2021-09-16', '2021-09-17', '2021-09-20',
           '2021-09-21', '2021-09-22', '2021-09-23', '2021-09-24', '2021-09-27', '2021-09-28', '2021-09-29', '2021-10-07',
           '2021-10-08', '2021-10-09', '2021-10-10', '2021-10-11', '2021-10-12', '2021-10-13', '2021-10-14', '2021-10-15',
           '2021-10-18', '2021-10-19', '2021-10-20', '2021-10-21', '2021-10-22', '2021-10-25', '2021-10-26', '2021-10-27',
           '2021-10-28', '2021-10-29', '2021-11-01', '2021-11-02', '2021-11-03', '2021-11-04', '2021-11-05', '2021-11-08',
           '2021-11-09', '2021-11-10', '2021-11-11', '2021-11-12', '2021-11-15', '2021-11-16', '2021-11-17', '2021-11-18',
           '2021-11-19', '2021-11-22', '2021-11-23', '2021-11-24', '2021-11-25', '2021-11-26', '2021-11-29', '2021-11-30',
           '2021-12-01', '2021-12-02', '2021-12-03', '2021-12-06', '2021-12-07', '2021-12-08', '2021-12-09', '2021-12-10',
           '2021-12-13', '2021-12-14', '2021-12-15', '2021-12-16', '2021-12-17', '2021-12-20', '2021-12-21', '2021-12-22',
           '2021-12-23', '2021-12-24', '2021-12-27', '2021-12-28', '2021-12-29', '2021-12-30', '2021-12-31']

#总方法
def main():
    user_name = input('请输入要统计的员工名字：')
    flag = selectName(user_name)
    if(flag == False):
        main()
    else:
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = '加班时长'
        sheet['A1'] = '姓名'
        sheet['B1'] = '日期'
        sheet['C1'] = '签到时间（起）'
        sheet['D1'] = '签到时间（止）'
        sheet['E1'] = '时长'
        sheet['F1'] = '合计（h）'
        l = len(flag)
        if len(flag):
            l1 = l + 1
            all_hours = 0
            for data in flag:
                name = data[0]
                date = data[1]
                ks_date = data[2]
                js_date = data[3]
                jbsc = data[4]
                all_hours += float(jbsc)
                sheet.append([name,date,ks_date,js_date,jbsc])
            sheet['F2'] = all_hours
            sheet.merge_cells('A2:A'+str(l1))
            sheet.merge_cells('F2:F'+str(l1))
            wb.save('D://'+user_name+'加班时长统计.xlsx')
            wb.close()
        else:
            print("这个人貌似没加班")
    flag = input('是否继续？1否,其他任意继续')
    if flag != '1':
        main()

#筛选数据
def selectName(user_name=''):
    list = []
    for i in range(5,max_row):
        obj = []
        name = sheet.cell(i, 1).value
        if name != user_name:
            continue

        all_date = sheet.cell(i, 5).value
        date = ''
        #week = ''
        if all_date.strip() != '':
            date = all_date[:10]
            #week = all_date[-3:]
        sb_time = sheet.cell(i,7).value
        sb_zt = sheet.cell(i,8).value
        if not sb_time:
            if sb_zt:
                if sb_zt.strip() == '未打卡':
                    sb_time = '08:30'
        xb_time = sheet.cell(i,9).value
        if bool(sb_time) and bool(xb_time):
            obj.append(name)
            obj.append(date)
            if date.strip() not in work_days:
                obj.append(sb_time)
            else:
                obj.append('17:30')
            obj.append(xb_time)
        else:
            continue
        list.append(obj)
    if len(list) == 0:
        print("查无此人")
        return False
    #print(list)
    user_list = calTime(list)
    return user_list

#计算加班时长
def calTime(user_list):
    return_list = []
    for user in user_list:
        if user[1].strip() in work_days:
            #工作日加班
            day1 = user[1].strip().split('-')
            day2 = user[1].strip().split('-')
            time = user[3].strip().split(':')
            day0 = user[2].strip().split(':')
            day1.extend(time)
            #加班过了凌晨
            today = datetime.datetime.strptime(user[1].strip(),'%Y-%m-%d')
            tomorrow = today + datetime.timedelta(days=1)
            tomo = str(tomorrow).split(' ')[0].split('-')
            time00 = ['8','30']
            day2.append('20')
            day2.append('00')

            tomorrow = int(day2[2]) + 1
            time1 = datetime.datetime(int(day1[0]),int(day1[1]),int(day1[2]),int(day1[3]),int(day1[4]))
            time11 = datetime.datetime(int(tomo[0]),int(tomo[1]),int(tomo[2]),int(day1[3]),int(day1[4]))
            time2 = datetime.datetime(int(day2[0]),int(day2[1]),int(day2[2]),int(day2[3]),int(day2[4]))
            time3 = datetime.datetime(int(tomo[0]),int(tomo[1]),int(tomo[2]),int(time00[0]),int(time00[1]))
            time0 = datetime.datetime(int(day2[0]),int(day2[1]),int(day2[2]),int(day0[0]),int(day0[1]))
            time000 = datetime.datetime(int(tomo[0]),int(tomo[1]),int(tomo[2]),int(0),int(0))
            if time1.__ge__(time2):
                #计算时间差 - 0.5小时
                seconds = (time1 - time0).seconds
                hours = seconds/3600 - 0.5
                jb_hours = math.floor(hours/0.5)*0.5
                #拼接加班时长
                user.append(jb_hours)
                return_list.append(user)

                #加班过了凌晨
            elif time11.__le__(time3):
                seconds = (time11 - time000).seconds
                hours = seconds/3600 + 6
                jb_hours = math.floor(hours/0.5)*0.5
                #拼接加班时长
                user.append(jb_hours)
                return_list.append(user)
            else:
                #为满足不计入统计
                continue
        else:
            #节假日加班
            day1 = user[1].strip().split('-')
            day2 = user[1].strip().split('-')
            time1 = user[2].strip().split(':')
            time2 = user[3].strip().split(':')
            day1.extend(time1)
            day2.extend(time2)
            a1 = datetime.datetime(int(day1[0]),int(day1[1]),int(day1[2]),int(day1[3]),int(day1[4]))
            a2 = datetime.datetime(int(day2[0]),int(day2[1]),int(day2[2]),int(day2[3]),int(day2[4]))

            time_sw = '11:30:00'
            time_xw = '13:00:00'

            sb = datetime.datetime.strptime(user[2].strip(),'%H:%M:%S')
            xb = datetime.datetime.strptime(user[3].strip(),'%H:%M:%S')
            sw = datetime.datetime.strptime(time_sw,'%H:%M:%S')
            xw = datetime.datetime.strptime(time_xw,'%H:%M:%S')

            #计算时间差
            hours = 0
            if sb < sw:
                seconds = (a2 - a1).seconds
                if xb < xw:
                    hours = seconds/3600
                else:
                    hours = seconds/3600 - 1.5
            elif sw <= sb < xw:
                a = datetime.datetime(int(day1[0]),int(day1[1]),int(day1[2]),int('13'),int('00'))
                seconds = (a2 - a).seconds
                hours = seconds/3600
            else:
                seconds = (a2 - a1).seconds
                hours = seconds/3600
            jb_hours = math.floor(hours/0.5)*0.5
            #拼接加班时长
            user.append(jb_hours)
            return_list.append(user)
    return return_list

try:
    global wb
    wb = openpyxl.load_workbook('E://技术中心考勤打卡日统计.xlsx')
except:
    print("未找到文件")
else:
    #获取所有的sheet页的名字
    sheetnames = wb.sheetnames
    #获取第一个sheet页
    sheet = wb[sheetnames[0]]
    #最大行数
    max_row = sheet.max_row + 1
    #最大列数
    #max_col = sheet.max_column + 1
    list = []
    main()

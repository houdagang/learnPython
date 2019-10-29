import openpyxl
wb = openpyxl.Workbook()
#wb.active就是获取这个工作薄的活动表，通常就是第一个工作表。
sheet = wb.active
#可以用.title给工作表重命名。现在第一个工作表的名称就会由原来默认的“sheet1”改为"new title"。
sheet.title = 'new title'
sheet['A1'] = '漫威'
row = ['美国队长','钢铁侠','蜘蛛侠']
#把我们想写入的一行内容写成列表，赋值给row。
sheet.append(row)
#用sheet.append()就能往表格里添加这一行文字。
wb.save('1.xlsx')

#读取
wb = openpyxl.load_workbook('1.xlsx')
sheet = wb['new title']
#获取所有的工作表，并打印
#sheetname = wb.sheetnames
#print(sheetname)

#A1_cell = sheet['A1']
#A1_value = A1_cell.value
#print(A1_value)

tit = []
data = []
row = sheet.max_row + 1
column = sheet.max_column + 1
for i in range(1,row):
    ele = {}
    for j in range(1,column):
        if i == 1:
            res = sheet.cell(i, j).value
            tit.append(res)
        else:
            ele[tit[j - 1]] = sheet.cell(i, j).value
    data.append(ele)

test_data = []
for item in data:
    if item != {}:
        test_data.append(item)
print(test_data)